import boto3
import subprocess
import openpyxl

# Initialize a session using Amazon EC2
ec2 = boto3.resource('ec2', region_name='ap-south-1')  # You can set your desired region

# AMI ID for Ubuntu 20.04 LTS in ap-south-1 (this may change; always verify the current ID)
ubuntu_ami_id = 'ami-0287a05f0ef0e9d9a'  # Replace with the actual AMI ID you find for your Ubuntu version and region

# User data script to update packages and install Nginx
password = 'lokesh'
user_data_script = f"""#!/bin/bash
cd /home/ubuntu/
sudo apt-get update -y
sudo apt-get install git -y
sudo git clone https://github.com/patilajayv/TravelMemory.git

# Installation instructions of NodeJS can be found here "https://github.com/nodesource/distributions.git"

sudo apt-get install -y ca-certificates curl gnupg
sudo mkdir -p /etc/apt/keyrings
curl -fsSL https://deb.nodesource.com/gpgkey/nodesource-repo.gpg.key | sudo gpg --dearmor -o /etc/apt/keyrings/nodesource.gpg  #Installation of NodeJS dependencies
NODE_MAJOR=20  # Current and Stable version of NodeJS
echo "deb [signed-by=/etc/apt/keyrings/nodesource.gpg] https://deb.nodesource.com/node_$NODE_MAJOR.x nodistro main" | sudo tee /etc/apt/sources.list.d/nodesource.list
sudo apt-get update
sudo apt-get install nodejs -y

cd /home/ubuntu/TravelMemory/frontend/
sudo npm install
sudo apt-get install nginx -y
sudo systemctl start nginx
sudo unlink /etc/nginx/sites-enabled/default
sudo cp /home/ubuntu/TravelMemory/mern-project /etc/nginx/sites-available/  # Reverse Proxy
sudo ln -s /etc/nginx/sites-available/mern-project /etc/nginx/sites-enabled/
sudo systemctl restart nginx 
sudo kill -9 $(sudo lsof -t -i:80)  #K Killing the port 80
npm start
"""


# Create a new EC2 instance
instances = ec2.create_instances(
    ImageId=ubuntu_ami_id,
    MinCount=1,
    MaxCount=1,
    InstanceType='t2.micro',
    KeyName='ubuntu_HVDevOps',  # Name of the key pair to use. Ensure it exists in your AWS account
    SecurityGroupIds=['sg-072870334bab90a65'],  # Ensure the security group exists and allows appropriate traffic
    UserData=user_data_script,  # Include the user data script
    TagSpecifications=[
        {
            'ResourceType': 'instance',
            'Tags': [
                {
                    'Key': 'Name',
                    'Value': 'ajay-lokeshFE'
                }
            ]
        }
    ]
)

instance = instances[0]
instance.wait_until_running()

# Open the existing Excel file and add a new worksheet if it doesn't exist and save the below details
try:
    workbook = openpyxl.load_workbook("instance_info_frontend.xlsx")
    worksheet = workbook.active
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Instance Info"
    worksheet['A1'] = "Instance ID"
    worksheet['B1'] = "Instance Type"
    worksheet['C1'] = "Public IPv4 address"
    worksheet['D1'] = "Private IP"
    worksheet['E1'] = "Launch time"
    worksheet['F1'] = "Security group name"

# Add instance information to the worksheet
next_row = worksheet.max_row + 1
worksheet.cell(row=next_row, column=1, value=instance.instance_id)
worksheet.cell(row=next_row, column=2, value=instance.instance_type)
worksheet.cell(row=next_row, column=3, value=instance.public_ip_address)
worksheet.cell(row=next_row, column=4, value=instance.private_ip_address)
worksheet.cell(row=next_row, column=5, value=str(instance.launch_time))
worksheet.cell(row=next_row, column=6, value=instance.security_groups[0]['GroupName'] if instance.security_groups else '')

# Save the Excel file
workbook.save("instance_info_frontend.xlsx")

# Wait for the instance to be in the running state
instance = instances[0]
instance.wait_until_running()

# Associate the Elastic IP address with the instance
ec2.meta.client.associate_address(InstanceId=instance.id, PublicIp='13.200.147.211')

print("New Ubuntu Frontend instance is created and associated with Elastic IP (13.200.147.211).")

# Run another Python script (boto3InstanceBE.py)
subprocess.run(['python', 'boto3InstanceBE.py'])
